"""Aptitude quiz routes."""
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from datetime import datetime
import random, io, base64
from openpyxl import load_workbook
from PIL import Image as PILImage
from api.config import get_db, APTITUDE_DIR

router = APIRouter()

GENERAL     = ["aptitude", "data-interpretation", "verbal-ability", "logical-reasoning", "verbal-reasoning"]
TECHNICAL   = ["c-programming", "cpp-programming", "c-sharp-programming", "java-programming"]

_cache: dict[str, list] = {}


def _extract_explanation(raw) -> str:
    if not raw:
        return "No explanation available."
    s = str(raw)
    idx = s.find("Explanation:")
    if idx == -1:
        return "No explanation available."
    extracted = s[idx + len("Explanation:"):].strip()
    if not extracted or "no answer description is available" in extracted.lower():
        return "No explanation available."
    return extracted


def _load(category: str) -> list:
    if category in _cache:
        return _cache[category]

    cats = GENERAL if category == "General" else TECHNICAL
    questions = []
    for sub in cats:
        path = APTITUDE_DIR / f"{sub}.xlsx"
        if not path.exists():
            continue
        wb = load_workbook(path)
        sheet = wb.active

        # Pre-index embedded images by row number
        img_by_row: dict[int, str] = {}
        try:
            for img in sheet._images:
                row_idx = img.anchor._from.row  # 0-based row index
                buf = io.BytesIO()
                PILImage.open(io.BytesIO(img._data())).save(buf, format="PNG")
                buf.seek(0)
                img_by_row[row_idx] = "data:image/png;base64," + base64.b64encode(buf.read()).decode()
        except Exception:
            pass

        for row in sheet.iter_rows(min_row=2):
            if len(row) < 5:
                continue
            q_no, q_text, options, answer, explanation = (row[i].value for i in range(5))
            if not (q_no and q_text and options and answer):
                continue

            # row[0].row is 1-based Excel row; image anchor is 0-based
            img_data = img_by_row.get(row[0].row - 1)

            sep = "\n" if sub == "non-verbal-reasoning" else ";"
            opts = [o.strip() for o in options.split(sep) if o.strip()]
            labeled = {chr(65 + i): o for i, o in enumerate(opts)}

            correct = str(answer).strip()
            for label, opt in labeled.items():
                if opt.strip().lower() == correct.lower():
                    correct = label
                    break

            questions.append({
                "id":              f"{sub}_{q_no}",
                "question_text":   str(q_text),
                "image_data":      img_data,
                "options":         opts,
                "labeled_options": labeled,
                "correct_answer":  correct,
                "explanation":     _extract_explanation(explanation),
            })

    _cache[category] = questions
    return questions


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/questions")
def get_questions(category: str = "General", count: int = 30):
    all_qs = _load(category)
    if not all_qs:
        raise HTTPException(404, f"No questions found for category '{category}'")
    return random.sample(all_qs, min(count, len(all_qs)))


class SubmitPayload(BaseModel):
    username: str
    category: str
    score:    int
    total:    int
    time_taken: float


@router.post("/submit")
def submit_test(p: SubmitPayload):
    col = get_db("quiz_system")["apti_test"]
    latest = list(col.find({"student_id": p.username, "category": p.category})
                      .sort("timestamp", -1).limit(1))
    test_no = latest[0]["test_no"] + 1 if latest else 1
    accuracy = round((p.score / p.total) * 100, 2) if p.total else 0.0

    past = list(col.find({"student_id": p.username, "category": p.category}))
    per_test: dict[int, list] = {}
    for rec in past:
        per_test.setdefault(rec["test_no"], []).append(rec)
    accs = []
    for recs in per_test.values():
        c   = sum(r.get("marks_achieved", 0) for r in recs)
        tot = sum(r.get("no_of_questions", 1) for r in recs)
        if tot:
            accs.append(round(c / tot * 100, 2))
    accs.append(accuracy)
    avg = round(sum(accs) / len(accs), 2)

    col.insert_one({
        "student_id":        p.username,
        "timestamp":         datetime.now(),
        "category":          p.category,
        "test_no":           test_no,
        "no_of_questions":   p.total,
        "marks_achieved":    p.score,
        "time_taken":        p.time_taken,
        "avg_test_accuracy": avg,
    })
    return {"test_no": test_no, "accuracy": accuracy, "avg_accuracy": avg}


class ViolationLog(BaseModel):
    username:  str
    violation: str


@router.post("/log-violation")
def log_violation(p: ViolationLog):
    get_db("quiz_system")["face_logs"].insert_one({
        "student_id": p.username,
        "timestamp":  datetime.now(),
        "violation":  p.violation,
    })
    return {"ok": True}


class TerminatePayload(BaseModel):
    username:        str
    category:        str
    violation_count: int
    violations:      list   # [{type, time}]


@router.post("/terminate")
def terminate_test(p: TerminatePayload):
    get_db("quiz_system")["cheating_reports"].insert_one({
        "student_id":      p.username,
        "test_type":       "aptitude",
        "category":        p.category,
        "violation_count": p.violation_count,
        "violations":      p.violations,
        "timestamp":       datetime.now(),
        "terminated":      True,
    })
    return {"ok": True}
